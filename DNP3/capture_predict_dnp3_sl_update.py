import os
import pandas as pd
from scapy.all import sniff, TCP, IP, Ether, Raw
from prediction_sl_update import preprocessed_data, load_pickle_files, get_predictions
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === DNP3 Function Code Names ===
DNP3_FUNC_NAMES = {
    0: "CONFIRM", 1: "READ", 2: "WRITE", 3: "SELECT",
    4: "OPERATE", 5: "DIRECT_OPERATE", 6: "DIRECT_OPERATE_NR",
    7: "IMMED_FREEZE", 8: "IMMED_FREEZE_NR", 9: "FREEZE_CLEAR",
    10: "FREEZE_CLEAR_NR", 11: "FREEZE_AT_TIME", 12: "FREEZE_AT_TIME_NR",
    13: "COLD_RESTART", 14: "WARM_RESTART", 15: "INITIALIZE_DATA",
    16: "INITIALIZE_APPL", 17: "START_APPL", 18: "STOP_APPL",
    19: "SAVE_CONFIG", 20: "ENABLE_UNSOLICITED", 21: "DISABLE_UNSOLICITED",
    22: "ASSIGN_CLASS", 23: "DELAY_MEASURE", 24: "RECORD_CURRENT_TIME",
    129: "RESPONSE", 130: "UNSOLICITED_RESPONSE",
}

# === Storage ===
dnp3_rows   = []
pkt_counter = 0
first_time  = None
last_time   = None
DNP3_PORT   = 20003  # ← change this one place only


# ── Packet Callback ────────────────────────────────────────────────────────────
def packet_callback(pkt):
    global pkt_counter, first_time, last_time, dnp3_rows

    if not pkt.haslayer(TCP) or not pkt.haslayer(IP):
        return

    tcp_layer = pkt[TCP]
    ip_layer  = pkt[IP]
    eth_layer = pkt[Ether] if pkt.haslayer(Ether) else None

    # Only DNP3 traffic
    if tcp_layer.sport != DNP3_PORT and tcp_layer.dport != DNP3_PORT:
        return

    # Must have raw payload
    if not pkt.haslayer(Raw):
        return

    raw = bytes(pkt[Raw])

    # DNP3 magic start bytes 0x05 0x64
    if len(raw) < 10 or raw[0] != 0x05 or raw[1] != 0x64:
        return

    pkt_counter += 1
    if first_time is None:
        first_time = float(pkt.time)

    frame_time          = float(pkt.time)
    frame_time_relative = frame_time - first_time
    frame_time_delta    = None if last_time is None else frame_time - last_time
    last_time           = frame_time

    # Parse DNP3 fields
    dnp3_len      = raw[2]
    dnp3_ctrl     = raw[3]
    dnp3_dst_addr = int.from_bytes(raw[4:6], 'little')
    dnp3_src_addr = int.from_bytes(raw[6:8], 'little')
    dnp3_func     = raw[12] if len(raw) > 12 else None
    dnp3_func_name = DNP3_FUNC_NAMES.get(dnp3_func, "UNKNOWN")

    # Control byte breakdown
    dnp3_dir       = (dnp3_ctrl & 0x80) >> 7
    dnp3_prm       = (dnp3_ctrl & 0x40) >> 6
    dnp3_func_code = dnp3_ctrl & 0x0F

    row = {
        # Frame
        "source_file"           : "live_capture",
        "frame.number"          : pkt_counter,
        "frame.time_epoch"      : frame_time,
        "frame.time_relative"   : frame_time_relative,
        "frame.time_delta"      : frame_time_delta,
        "frame.len"             : len(pkt),
        "frame.cap_len"         : len(bytes(pkt)),

        # Ethernet
        "eth.src"               : eth_layer.src if eth_layer else None,
        "eth.dst"               : eth_layer.dst if eth_layer else None,

        # IP
        "ip.src"                : ip_layer.src,
        "ip.dst"                : ip_layer.dst,

        # TCP
        "tcp.srcport"           : tcp_layer.sport,
        "tcp.dstport"           : tcp_layer.dport,
        "tcp.flags"             : tcp_layer.flags.value,
        "tcp.len"               : len(tcp_layer.payload),
        "tcp.window_size_value" : tcp_layer.window,
        "tcp.time_delta"        : frame_time_delta,

        # DNP3
        "dnp3.len"              : dnp3_len,
        "dnp3.ctrl"             : dnp3_ctrl,
        "dnp3.dst_addr"         : dnp3_dst_addr,
        "dnp3.src_addr"         : dnp3_src_addr,
        "dnp3.dir"              : dnp3_dir,
        "dnp3.prm"              : dnp3_prm,
        "dnp3.func_code_link"   : dnp3_func_code,
        "dnp3.func_code"        : dnp3_func,
        "dnp3.func_name"        : dnp3_func_name,
        "dnp3.payload_len"      : len(raw),
    }

    dnp3_rows.append(row)
    print(f"[{pkt_counter}] {ip_layer.src} → {ip_layer.dst} | {dnp3_func_name}")


# ==== Main ====
if __name__ == "__main__":
 
    # ── Choose mode here ─────────────────────────────────────────────────────
    MODE  = "multiclass"   # "binary" or "multiclass"
    # ─────────────────────────────────────────────────────────────────────────
 
    iface = r"\Device\NPF_Loopback"
    PATH  = "../DNP3/Model_Files"
 
    # Pickle filenames differ per mode
    if MODE == "binary":
        MODEL_FILE  = "DNP3_RF_best_model (1).pkl"
        SCALER_FILE = "dnp3_scaler_for_SL (1).pkl"
        OUTPUT_FILE = "dnp3_capture_with_predictions_SL_binary.xlsx"
    else:
        MODEL_FILE  = "DNP3_RF_best_model__RFFF.pkl"
        SCALER_FILE = "dnp3_scaler_for_SL_RF.pkl"
        OUTPUT_FILE = "dnp3_capture_with_predictions_SL_multiclass.xlsx"
 
    print("=" * 60)
    print("  DNP3 Live Capture + Window-Based Prediction")
    print(f"  Mode      : {MODE.upper()}")
    print(f"  Interface : {iface}")
    print(f"  Filter    : tcp port {DNP3_PORT}")
    print(f"  Timeout   : 30 seconds")
    print(f"  Window    : 10 packets (sliding)")
    print("=" * 60)
    print("  Waiting for packets... (run master.py now)")
    print()
 
    sniff(
        prn=packet_callback,
        filter=f"tcp port {DNP3_PORT}",
        iface=iface,
        store=False,
        timeout=30
    )
 
    print(f"\n✅ Captured {len(dnp3_rows)} DNP3 packets")
 
    if len(dnp3_rows) == 0:
        print("⚠️  No DNP3 packets captured.")
        print("    Check: outstation running? attack generator running?")
        exit()
 
    # ── Step 1: Save raw capture ──
    df_raw = pd.DataFrame(dnp3_rows)
    df_raw.to_csv("dnp3_capture.csv", index=False)
    print(f"📂 Raw capture saved → dnp3_capture.csv  ({len(df_raw)} rows)")
 
    # ── Step 2: Preprocess (same for both modes) ──
    df_proc = preprocessed_data(df_raw)
 
    if len(df_proc) < 10:
        print(f"⚠️  Only {len(df_proc)} packets after filtering — need at least 10 for windowed prediction.")
        exit()
 
    # ── Step 3: Load model + scaler ──
    model, scaler = load_pickle_files(
        path=PATH,
        model_filename=MODEL_FILE,
        scaler_filename=SCALER_FILE
    )
 
    # ── Step 4: Predict ──
    labels, proba_df = get_predictions(
        df=df_proc,
        model=model,
        scaler=scaler,
        mode=MODE,
        window_size=10
    )
 
    # ── Step 5: Summary ──
    print("\n📊 Prediction Summary:")
    print("-" * 40)
    from collections import Counter
    valid_labels = [l for l in labels if l is not None]
    for label, cnt in Counter(valid_labels).items():
        bar = "█" * min(cnt, 40)
        print(f"  {label:25s} : {cnt:4d}  {bar}")
    print("-" * 40)
    print(f"  Total packets   : {len(labels)}")
    print(f"  Predicted       : {len(valid_labels)}")
    print(f"  Waiting (buffer): {labels.count(None)}")
 
    # ── Step 6: Combine + save Excel ──
    df_proc  = df_proc.reset_index(drop=True)
    proba_df = proba_df.reset_index(drop=True)
    df_proc["predicted_label"] = labels
 
    df_out = pd.concat([df_proc, proba_df], axis=1)
    df_out.to_excel(OUTPUT_FILE, index=False)
 
    # Highlight attack rows in red
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    red_fill  = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    col_index = list(df_out.columns).index("predicted_label") + 1
 
    for row in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=col_index).value
        if cell_val not in (None, "NORMAL", "WAITING"):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = red_fill
 
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Saved → {OUTPUT_FILE}")