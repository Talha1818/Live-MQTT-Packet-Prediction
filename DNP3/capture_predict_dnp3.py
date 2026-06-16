import os
import pandas as pd
from scapy.all import sniff, TCP, IP, Ether, Raw
from prediction import preprocessed_data, load_pickle_files, get_predictions
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === DNP3 Function Code Names ===
DNP3_FUNC_NAMES = {
    0: "CONFIRM", 1: "READ", 2: "WRITE", 3: "SELECT",
    4: "OPERATE", 5: "DIRECT_OPERATE", 6: "DIRECT_OPERATE_NR",
    7: "IMMED_FREEZE", 8: "IMMED_FREEZE_NR", 9: "FREEZE_CLEAR",
    10: "FREEZE_CLEAR_NR", 11: "FREEZE_AT_TIME", 12: "FREEZE_AT_TIME_NR",
    13: "COLD_RESTART", 14: "WARM_RESTART", 15: "INITIALIZE_DATA",
    16: "INITIALIZE_APPL", 17: "START_APPL", 18: "STOP_APPL",
    19: "SAVE_CONFIG", 20: "ENABLE_UNSOLICITED", 21: "DISABLE_UNSOLICITED",
    22: "ASSIGN_CLASS", 23: "DELAY_MEASURE", 24: "RECORD_CURRENT_TIME",
    129: "RESPONSE", 130: "UNSOLICITED_RESPONSE",
}

# === Storage ===
dnp3_rows  = []
pkt_counter = 0
first_time, last_time = None, None

# ── Packet Callback (same logic as parse_pcap_files) ──────────────────────────
def packet_callback(pkt):
    global pkt_counter, first_time, last_time, dnp3_rows

    if not pkt.haslayer(TCP) or not pkt.haslayer(IP):
        return

    tcp_layer = pkt[TCP]
    ip_layer  = pkt[IP]
    eth_layer = pkt[Ether] if pkt.haslayer(Ether) else None

    # ── Only DNP3 traffic (port 20003) ──
    if tcp_layer.sport != 20003 and tcp_layer.dport != 20003:
        return

    # ── Must have raw payload ──
    if not pkt.haslayer(Raw):
        return

    raw = bytes(pkt[Raw])

    # ── DNP3 magic start bytes 0x05 0x64 ──
    if len(raw) < 10 or raw[0] != 0x05 or raw[1] != 0x64:
        return

    pkt_counter += 1
    if first_time is None:
        first_time = float(pkt.time)

    frame_time          = float(pkt.time)
    frame_time_relative = frame_time - first_time
    frame_time_delta    = None if last_time is None else frame_time - last_time
    last_time           = frame_time

    # ── Parse DNP3 fields ──
    dnp3_len       = raw[2]
    dnp3_ctrl      = raw[3]
    dnp3_dst_addr  = int.from_bytes(raw[4:6], 'little')
    dnp3_src_addr  = int.from_bytes(raw[6:8], 'little')
    dnp3_func      = raw[12] if len(raw) > 12 else None
    dnp3_func_name = DNP3_FUNC_NAMES.get(dnp3_func, "UNKNOWN")

    # ── Control byte breakdown ──
    dnp3_dir       = (dnp3_ctrl & 0x80) >> 7
    dnp3_prm       = (dnp3_ctrl & 0x40) >> 6
    dnp3_fcb       = (dnp3_ctrl & 0x20) >> 5
    dnp3_fcv       = (dnp3_ctrl & 0x10) >> 4
    dnp3_func_code = dnp3_ctrl & 0x0F

    row = {
        # === Frame ===
        "source_file"           : "live_capture",
        "frame.number"          : pkt_counter,
        "frame.time_epoch"      : frame_time,
        "frame.time_relative"   : frame_time_relative,
        "frame.time_delta"      : frame_time_delta,
        "frame.len"             : len(pkt),
        "frame.cap_len"         : len(bytes(pkt)),

        # === Ethernet ===
        "eth.src"               : eth_layer.src if eth_layer else None,
        "eth.dst"               : eth_layer.dst if eth_layer else None,

        # === IP ===
        "ip.src"                : ip_layer.src,
        "ip.dst"                : ip_layer.dst,
        "ip.proto"              : ip_layer.proto,
        "ip.ttl"                : ip_layer.ttl,

        # === TCP ===
        "tcp.srcport"           : tcp_layer.sport,
        "tcp.dstport"           : tcp_layer.dport,
        "tcp.flags"             : tcp_layer.flags.value,
        "tcp.len"               : len(tcp_layer.payload),
        "tcp.window_size_value" : tcp_layer.window,
        "tcp.checksum"          : tcp_layer.chksum,
        "tcp.time_delta"        : frame_time_delta,

        # === DNP3 Data Link Layer ===
        "dnp3_present"          : True,
        "dnp3.start"            : f"0x{raw[0]:02X}{raw[1]:02X}",
        "dnp3.len"              : dnp3_len,
        "dnp3.ctrl"             : dnp3_ctrl,
        "dnp3.dst_addr"         : dnp3_dst_addr,
        "dnp3.src_addr"         : dnp3_src_addr,
        "dnp3.dir"              : dnp3_dir,
        "dnp3.prm"              : dnp3_prm,
        "dnp3.fcb"              : dnp3_fcb,
        "dnp3.fcv"              : dnp3_fcv,
        "dnp3.func_code_link"   : dnp3_func_code,

        # === DNP3 Application Layer ===
        "dnp3.func_code"        : dnp3_func,
        "dnp3.func_name"        : dnp3_func_name,
        "dnp3.payload_len"      : len(raw),
        "dnp3.raw_hex"          : raw.hex()[:64],
    }

    dnp3_rows.append(row)
    print(f"[{pkt_counter}] {ip_layer.src} → {ip_layer.dst} | {dnp3_func_name}")


if __name__ == "__main__":

    iface = r"\Device\NPF_Loopback"
    PATH  = "../DNP3/Model_Files"

    # ── Step 1: Live Capture ──
    print("🚦 Capturing live DNP3 packets on ports 20000-20003...")
    sniff(prn=packet_callback, filter="tcp portrange 20000-20003",
          iface=iface, store=False, timeout=30)
    print(f"\n✅ Captured {len(dnp3_rows)} DNP3 packets")

    if len(dnp3_rows) == 0:
        print("⚠️ No DNP3 packets captured. Check interface or simulator.")
        exit()

    # ── Step 2: Save raw capture ──
    df = pd.DataFrame(dnp3_rows)
    df = df[df["dnp3.func_code"].notna()].copy()
    df.to_csv("dnp3_capture.csv", index=False)
    print("📂 Saved to dnp3_capture.csv")

    # ── Step 3: Preprocess + Predict ──
    df_original, df_proc = preprocessed_data(df)
    model, scaler        = load_pickle_files(PATH)
    labels, proba_df      = get_predictions(df_proc, model, scaler)

    df_original["predicted_label"] = labels
    df_original = df_original.reset_index(drop=True)
    proba_df    = proba_df.reset_index(drop=True)
    df_original = pd.concat([df_original, proba_df], axis=1)

    # ── Step 4: Save Excel with attack highlighting ──
    file_name = "dnp3_capture_with_predictions.xlsx"
    df_original.to_excel(file_name, index=False)

    wb = load_workbook(file_name)
    ws = wb.active
    red_fill  = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    col_index = list(df_original.columns).index("predicted_label") + 1

    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=col_index).value != "NORMAL":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = red_fill

    wb.save(file_name)
    print("✅ Predictions with highlighted attacks saved to Excel")